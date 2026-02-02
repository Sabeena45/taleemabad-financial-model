#!/usr/bin/env python3
"""
Portfolio Wins Tracker Generator
Creates an Excel file with progress tracking, charts, and conditional formatting
"""

import os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.formatting.rule import FormulaRule, DataBarRule
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("Installing openpyxl...")
    os.system("pip3 install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.formatting.rule import FormulaRule, DataBarRule
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

# === DATA ===
WINS = {
    "FUNDRAISING": [
        ("Anchor funder secured", "Multi-year partner who funds the model, not an activity", "High"),
        ("Cost per child story locked", "A number you can say in one breath, with proof, that goes down over time", "High"),
        ("Pipeline not panic", "6-9 months of confirmed or near-confirmed runway visible", "High"),
        ("Inbound interest", "At least 2 serious funders or networks reaching out to you", "Medium"),
        ("Flagship narrative deck", "The Taleemabad story - tight, data-led, reusable everywhere", "High"),
    ],
    "POLICY": [
        ("Government repeats your language", "They say 'AI coaching', 'fidelity', 'continuous observation' - your concepts", "High"),
        ("Model in government plan", "Written into a plan or notification - system language, not pilot talk", "High"),
        ("Invited to policy tables", "Advisory roles, working groups, consultations - not pitching for them", "Medium"),
        ("Project → System shift", "Conversations move from 'your program' to 'how the system should work'", "Medium"),
        ("Senior champion inside", "Someone who defends the model when you're not in the room", "High"),
    ],
    "IMPACT": [
        ("Clean learning result", "One undeniable number that survives scrutiny", "High"),
        ("Fidelity drives decisions", "Teams change behavior because of dashboards, not opinions", "Medium"),
        ("Predictive capability", "'If X teachers use it this way, we expect Y improvement'", "Medium"),
        ("External validation", "Research partner, evaluator, publication, or working paper", "High"),
        ("Impact closes money", "Data directly closes funding deals", "High"),
    ],
}

# === COLORS ===
COLORS = {
    "header": "1F2937",      # Dark gray
    "fundraising": "10B981", # Green
    "policy": "3B82F6",      # Blue
    "impact": "F59E0B",      # Amber
    "done": "D1FAE5",        # Light green
    "in_progress": "FEF3C7", # Light yellow
    "not_started": "F3F4F6", # Light gray
    "white": "FFFFFF",
}

def create_tracker():
    wb = Workbook()

    # === TRACKER SHEET ===
    ws = wb.active
    ws.title = "Tracker"

    # Column widths
    col_widths = [14, 35, 14, 10, 12, 30, 40, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Header row
    headers = ["Category", "Win", "Status", "Priority", "Target", "Evidence", "Notes", "Updated"]
    header_fill = PatternFill(start_color=COLORS["header"], end_color=COLORS["header"], fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 25

    # Status dropdown validation
    status_dv = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Done"',
        allow_blank=True
    )
    status_dv.error = "Please select from dropdown"
    status_dv.errorTitle = "Invalid Status"
    ws.add_data_validation(status_dv)

    # Priority dropdown validation
    priority_dv = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=True
    )
    ws.add_data_validation(priority_dv)

    # Add data rows
    row = 2
    category_colors = {
        "FUNDRAISING": COLORS["fundraising"],
        "POLICY": COLORS["policy"],
        "IMPACT": COLORS["impact"],
    }

    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )

    for category, wins in WINS.items():
        cat_fill = PatternFill(start_color=category_colors[category], end_color=category_colors[category], fill_type="solid")

        for win_name, win_desc, priority in wins:
            ws.cell(row=row, column=1, value=category).fill = cat_fill
            ws.cell(row=row, column=1).font = Font(color="FFFFFF", bold=True, size=10)
            ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")

            ws.cell(row=row, column=2, value=win_name)
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="center")

            status_cell = ws.cell(row=row, column=3, value="Not Started")
            status_dv.add(status_cell)
            status_cell.alignment = Alignment(horizontal="center", vertical="center")

            priority_cell = ws.cell(row=row, column=4, value=priority)
            priority_dv.add(priority_cell)
            priority_cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.cell(row=row, column=5, value="")  # Target date
            ws.cell(row=row, column=6, value="")  # Evidence
            ws.cell(row=row, column=7, value=win_desc)  # Notes (pre-filled with description)
            ws.cell(row=row, column=7).alignment = Alignment(wrap_text=True, vertical="center")
            ws.cell(row=row, column=8, value="")  # Updated

            # Apply borders
            for col in range(1, 9):
                ws.cell(row=row, column=col).border = thin_border

            ws.row_dimensions[row].height = 35
            row += 1

        # Add empty row for spacing between categories
        row += 1

    # Add conditional formatting for status column
    done_fill = PatternFill(start_color=COLORS["done"], end_color=COLORS["done"], fill_type="solid")
    progress_fill = PatternFill(start_color=COLORS["in_progress"], end_color=COLORS["in_progress"], fill_type="solid")

    # === DASHBOARD SHEET ===
    ds = wb.create_sheet("Dashboard")

    # Title
    ds.merge_cells('A1:F1')
    ds['A1'] = "PORTFOLIO WINS DASHBOARD"
    ds['A1'].font = Font(size=20, bold=True, color=COLORS["header"])
    ds['A1'].alignment = Alignment(horizontal="center")
    ds.row_dimensions[1].height = 40

    # Last updated
    ds['A2'] = f"Last Updated: {datetime.now().strftime('%B %d, %Y')}"
    ds['A2'].font = Font(size=10, italic=True, color="6B7280")

    # Summary section
    ds['A4'] = "PROGRESS SUMMARY"
    ds['A4'].font = Font(size=14, bold=True)

    # Category headers
    ds['A6'] = "Category"
    ds['B6'] = "Done"
    ds['C6'] = "In Progress"
    ds['D6'] = "Not Started"
    ds['E6'] = "Total"
    ds['F6'] = "% Complete"

    for col in range(1, 7):
        ds.cell(row=6, column=col).font = Font(bold=True)
        ds.cell(row=6, column=col).fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    # Formulas for counting (these reference the Tracker sheet)
    categories = ["FUNDRAISING", "POLICY", "IMPACT"]
    for i, cat in enumerate(categories, 7):
        ds.cell(row=i, column=1, value=cat)
        ds.cell(row=i, column=1).font = Font(bold=True)

        # Count formulas
        ds.cell(row=i, column=2, value=f'=COUNTIFS(Tracker!A:A,"{cat}",Tracker!C:C,"Done")')
        ds.cell(row=i, column=3, value=f'=COUNTIFS(Tracker!A:A,"{cat}",Tracker!C:C,"In Progress")')
        ds.cell(row=i, column=4, value=f'=COUNTIFS(Tracker!A:A,"{cat}",Tracker!C:C,"Not Started")')
        ds.cell(row=i, column=5, value=f'=COUNTIF(Tracker!A:A,"{cat}")')
        ds.cell(row=i, column=6, value=f'=IF(E{i}>0,B{i}/E{i},0)')
        ds.cell(row=i, column=6).number_format = '0%'

    # Totals row
    ds['A10'] = "TOTAL"
    ds['A10'].font = Font(bold=True)
    ds['B10'] = '=SUM(B7:B9)'
    ds['C10'] = '=SUM(C7:C9)'
    ds['D10'] = '=SUM(D7:D9)'
    ds['E10'] = '=SUM(E7:E9)'
    ds['F10'] = '=IF(E10>0,B10/E10,0)'
    ds['F10'].number_format = '0%'

    # Column widths for dashboard
    ds.column_dimensions['A'].width = 15
    ds.column_dimensions['B'].width = 12
    ds.column_dimensions['C'].width = 12
    ds.column_dimensions['D'].width = 12
    ds.column_dimensions['E'].width = 10
    ds.column_dimensions['F'].width = 12

    # Add a bar chart
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.title = "Progress by Category"
    chart.style = 10

    # Data for chart (Done, In Progress, Not Started)
    data = Reference(ds, min_col=2, min_row=6, max_col=4, max_row=9)
    cats = Reference(ds, min_col=1, min_row=7, max_row=9)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 4
    chart.width = 15
    chart.height = 10

    ds.add_chart(chart, "A13")

    # === INSTRUCTIONS SHEET ===
    ins = wb.create_sheet("Instructions")

    instructions = [
        ("HOW TO USE THIS TRACKER", ""),
        ("", ""),
        ("1. UPDATING STATUS", "Go to Tracker sheet → Click Status cell → Select from dropdown"),
        ("", "Options: Not Started, In Progress, Done"),
        ("", ""),
        ("2. ADDING NEW ITEMS", "Insert a new row in the Tracker sheet"),
        ("", "Copy the format from an existing row"),
        ("", "Make sure to set the Category correctly"),
        ("", ""),
        ("3. VIEWING PROGRESS", "Go to Dashboard sheet"),
        ("", "Charts and percentages update automatically"),
        ("", ""),
        ("4. ADDING EVIDENCE", "Use the Evidence column for links to documents, data, etc."),
        ("", ""),
        ("5. SYNCING WITH CLAUDE", "Export this sheet as CSV periodically"),
        ("", "Share updates in conversation"),
        ("", "Or paste the Tracker sheet contents"),
        ("", ""),
        ("TIPS", ""),
        ("", "- Update the 'Updated' column when you make changes"),
        ("", "- Use Notes for context and next steps"),
        ("", "- High priority items should be tackled first"),
        ("", "- Review weekly with Claude for accountability"),
    ]

    for i, (title, content) in enumerate(instructions, 1):
        if title:
            ins.cell(row=i, column=1, value=title).font = Font(bold=True, size=12)
        if content:
            ins.cell(row=i, column=2, value=content)

    ins.column_dimensions['A'].width = 25
    ins.column_dimensions['B'].width = 60

    # Save
    output_path = os.path.join(os.path.dirname(__file__), "Portfolio_Wins_Tracker.xlsx")
    wb.save(output_path)
    print(f"Created: {output_path}")
    return output_path

if __name__ == "__main__":
    create_tracker()
