"""
Excel Export for Taleemabad Financial Model.
Generates multi-sheet workbook with all financial data.
"""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
import pandas as pd


def export_to_excel(
    cashflow_model,
    scenario_model,
    sensitivity_model,
    custom_assumptions: dict,
) -> bytes:
    """
    Generate Excel workbook with all financial data.

    Args:
        cashflow_model: CashFlowModel instance
        scenario_model: ScenarioModel instance
        sensitivity_model: SensitivityModel instance
        custom_assumptions: Dict of user-modified assumptions

    Returns:
        bytes: Excel file as bytes
    """
    output = io.BytesIO()
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    currency_format = '"$"#,##0'
    percent_format = '0.0%'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # =========================================================================
    # Sheet 1: Executive Summary
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Executive Summary"

    # Title
    ws1['A1'] = "Taleemabad Financial Model - 2026"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1.merge_cells('A1:D1')

    ws1['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws1['A2'].font = Font(italic=True, color="666666")

    # Key metrics
    ws1['A4'] = "Key Metrics"
    ws1['A4'].font = Font(bold=True, size=14)

    metrics = [
        ("Metric", "Value"),
        ("Opening Balance (Jan 1, 2026)", custom_assumptions.get('opening_balance', 723248)),
        ("Total Inflows", cashflow_model.get_total_inflows()),
        ("Total Expenses", cashflow_model.get_total_outflows()),
        ("Net Cash Flow", cashflow_model.get_net_cash_flow()),
        ("Year-End Surplus", cashflow_model.get_year_end_position()),
        ("Exchange Rate (PKR/USD)", custom_assumptions.get('exchange_rate', 283)),
        ("Runway (Months)", round(cashflow_model.get_year_end_position() / cashflow_model.get_average_monthly_burn(), 1)),
    ]

    for row_idx, (metric, value) in enumerate(metrics, start=5):
        ws1.cell(row=row_idx, column=1, value=metric)
        if row_idx == 5:  # Header row
            ws1.cell(row=row_idx, column=1).font = header_font
            ws1.cell(row=row_idx, column=1).fill = header_fill
            ws1.cell(row=row_idx, column=2).font = header_font
            ws1.cell(row=row_idx, column=2).fill = header_fill

        cell = ws1.cell(row=row_idx, column=2, value=value)
        if isinstance(value, (int, float)) and metric not in ["Exchange Rate (PKR/USD)", "Runway (Months)"]:
            cell.number_format = currency_format

    # Adjust column widths
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 20

    # =========================================================================
    # Sheet 2: Monthly Cash Flow
    # =========================================================================
    ws2 = wb.create_sheet("Cash Flow")

    ws2['A1'] = "Monthly Cash Flow - 2026"
    ws2['A1'].font = Font(bold=True, size=14)

    # Get cash flow data
    cf_data = cashflow_model.to_dataframe_dict()
    df_cf = pd.DataFrame(cf_data)

    # Headers
    headers = ["Month", "Opening", "Inflows", "Outflows", "Net", "Closing"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws2.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    cumulative = custom_assumptions.get('opening_balance', 723248)
    for row_idx, row_data in enumerate(cf_data, start=4):
        ws2.cell(row=row_idx, column=1, value=row_data['Month'])
        ws2.cell(row=row_idx, column=2, value=cumulative).number_format = currency_format
        ws2.cell(row=row_idx, column=3, value=row_data['Inflows']).number_format = currency_format
        ws2.cell(row=row_idx, column=4, value=row_data['Outflows']).number_format = currency_format
        net = row_data['Inflows'] - row_data['Outflows']
        ws2.cell(row=row_idx, column=5, value=net).number_format = currency_format
        cumulative += net
        ws2.cell(row=row_idx, column=6, value=cumulative).number_format = currency_format

    # Totals row
    total_row = len(cf_data) + 4
    ws2.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=total_row, column=3, value=cashflow_model.get_total_inflows()).number_format = currency_format
    ws2.cell(row=total_row, column=4, value=cashflow_model.get_total_outflows()).number_format = currency_format
    ws2.cell(row=total_row, column=5, value=cashflow_model.get_net_cash_flow()).number_format = currency_format

    # Column widths
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws2.column_dimensions[col].width = 15

    # =========================================================================
    # Sheet 3: Scenario Comparison
    # =========================================================================
    ws3 = wb.create_sheet("Scenarios")

    ws3['A1'] = "Scenario Analysis"
    ws3['A1'].font = Font(bold=True, size=14)

    # Run scenarios
    scenario_model.run_all_scenarios()
    comparison = scenario_model.compare_scenarios()

    # Headers
    scenario_headers = ["Scenario", "Total Inflows", "Total Expenses", "Year-End Surplus", "Runway (Months)", "Minimum Cash"]
    for col_idx, header in enumerate(scenario_headers, start=1):
        cell = ws3.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Data
    for row_idx, (scenario_name, data) in enumerate(comparison.items(), start=4):
        ws3.cell(row=row_idx, column=1, value=scenario_name)
        ws3.cell(row=row_idx, column=2, value=data['total_inflows']).number_format = currency_format
        ws3.cell(row=row_idx, column=3, value=data['total_expenses']).number_format = currency_format
        ws3.cell(row=row_idx, column=4, value=data['year_end_surplus']).number_format = currency_format
        ws3.cell(row=row_idx, column=5, value=round(data['runway_months'], 1))
        ws3.cell(row=row_idx, column=6, value=data['minimum_cash']).number_format = currency_format

    # Column widths
    ws3.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws3.column_dimensions[col].width = 18

    # =========================================================================
    # Sheet 4: Grant Analysis
    # =========================================================================
    ws4 = wb.create_sheet("Grant Risk")

    ws4['A1'] = "Grant Dependency Analysis"
    ws4['A1'].font = Font(bold=True, size=14)

    grant_analysis = sensitivity_model.analyze_grant_dependency()

    # Headers
    grant_headers = ["Grant", "Amount", "% of Total", "Impact if Lost", "New Surplus", "Critical?"]
    for col_idx, header in enumerate(grant_headers, start=1):
        cell = ws4.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Data
    for row_idx, (grant_name, data) in enumerate(grant_analysis.items(), start=4):
        ws4.cell(row=row_idx, column=1, value=grant_name.replace('_', ' ').title())
        ws4.cell(row=row_idx, column=2, value=data['grant_amount']).number_format = currency_format
        ws4.cell(row=row_idx, column=3, value=data['percentage_of_total'] / 100).number_format = percent_format
        ws4.cell(row=row_idx, column=4, value=data['impact_on_surplus']).number_format = currency_format
        ws4.cell(row=row_idx, column=5, value=data['new_surplus']).number_format = currency_format
        ws4.cell(row=row_idx, column=6, value="Yes" if data['critical'] else "No")

        # Highlight critical grants
        if data['critical']:
            for col in range(1, 7):
                ws4.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
                )

    # Column widths
    ws4.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws4.column_dimensions[col].width = 15

    # =========================================================================
    # Sheet 5: Assumptions
    # =========================================================================
    ws5 = wb.create_sheet("Assumptions")

    ws5['A1'] = "Model Assumptions"
    ws5['A1'].font = Font(bold=True, size=14)

    assumptions_data = [
        ("Parameter", "Value", "Notes"),
        ("Opening Balance", custom_assumptions.get('opening_balance', 723248), "As of Jan 1, 2026"),
        ("Exchange Rate", custom_assumptions.get('exchange_rate', 283), "PKR/USD"),
        ("Expense Multiplier", custom_assumptions.get('expense_multiplier', 1.0), "1.0 = baseline"),
        ("Data Source", "Budget 2026 v2.0", "Draft Baseline Internal"),
    ]

    for row_idx, row_data in enumerate(assumptions_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws5.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:
                cell.font = header_font
                cell.fill = header_fill

    # Column widths
    ws5.column_dimensions['A'].width = 20
    ws5.column_dimensions['B'].width = 15
    ws5.column_dimensions['C'].width = 30

    # Save
    wb.save(output)
    return output.getvalue()
